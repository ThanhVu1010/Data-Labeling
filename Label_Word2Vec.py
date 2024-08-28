import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import logging
from gensim.models import Word2Vec
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np

# Thiết lập logging
logging.basicConfig(level=logging.INFO)

# Tên cột và sheet (Nhớ để ý phần này trước khi chạy)
import_file_paths = [
    '/content/drive/MyDrive/label/im_hs 50_2022_converted_date.xlsx'
]
keywords_file_path = '/content/drive/MyDrive/label/keyword.xlsx'

# Tên cột và sheet (Nhớ để ý phần này trước khi chạy)
sheet_name_supply_chain = 'Supply Chain_IM'
sheet_name_product = 'Product_IM'
sheet_name_material = 'Structure_IM'
sheet_name_model_pattern = 'Form_IM'
sheet_name_function = 'Function_IM'
sheet_name_shape = 'Shape_IM'
sheet_name_pattern = 'Pattern_IM'

col_keyword = 'Keyword'
col_products = 'Products'
col_corrected_products = 'Corrected_Products'
col_supply_chain = 'Chuỗi cung ứng'
col_product_label = 'Tên sản phẩm'
col_yarn_count = 'Chi số sợi'
col_material_name = 'Phân bổ thành phần'
col_material_name_no_percentage = 'Tên nguyên liệu'
col_model_pattern = 'Kiểu mẫu'
col_function = 'Chức năng'
col_shape = 'Hình dáng'
col_pattern = 'Họa tiết'
col_Thickness = 'Độ dày'
col_Width = 'Độ rộng'
col_Weight_range = 'Trọng lượng'
col_fabric_size = 'Khổ vải'

# Load data from Excel
def load_data(file_path, sheet_name=None):
    try:
        return pd.read_excel(file_path, sheet_name=sheet_name if sheet_name else 'Sheet1')
    except FileNotFoundError:
        logging.error(f"File not found: {file_path}")
        return pd.DataFrame()
    except Exception as e:
        logging.error(f"Error loading data from {file_path}: {e}")
        return pd.DataFrame()

# Correct spelling function
correction_dict = {'tămg': 'tằm', 'tăm': 'tằm', 'tầm': 'tằm'}

def correct_spelling_with_dict(product):
    if pd.isnull(product):
        return ''

    cleaned_product = str(product).replace('_', ' ')
    pattern = re.compile(r'(\b[\d]+\.\d+\b|\b[\w]+\b|/|[A-Za-z]+\d+|\d+[A-Za-z]+|\b\w+\.\w+|\w+\b)')
    words = pattern.findall(cleaned_product.title())
    corrected_words = [correction_dict.get(word.lower(), word) for word in words]
    corrected_text = ' '.join(corrected_words)

    return corrected_text

# Train Word2Vec model on product descriptions
def train_word2vec(sentences):
    """
    Trains a Word2Vec model on a list of sentences.
    """
    model = Word2Vec(sentences, vector_size=100, window=5, min_count=1, workers=4)
    return model

# Convert a text to a vector using Word2Vec
def vectorize_text(text, model):
    words = text.lower().split()
    word_vectors = [model.wv[word] for word in words if word in model.wv]
    if not word_vectors:  # Nếu không có từ nào có trong mô hình
        return np.zeros((model.vector_size,))
    return np.mean(word_vectors, axis=0)

# Label text based on Word2Vec similarity with a check for keyword presence
def label_using_word2vec(text, keywords, model, output_column, supply_chain=None):
    text = text.lower()
    text_vector = vectorize_text(text, model)
    max_similarity = -1
    best_keyword = 'No keyword matched'
    best_label = 'No label found'

    # Quy tắc đặc biệt cho "Kiểu mẫu" khi chuỗi cung ứng là "vải"
    if output_column == col_model_pattern and supply_chain and supply_chain.lower() == "vải":
        return determine_material_type(text), best_keyword

    for _, row in keywords.iterrows():
        keyword = str(row[col_keyword]).lower()
        if keyword in text:
            keyword_vector = vectorize_text(row[col_keyword], model)
            similarity = cosine_similarity([text_vector], [keyword_vector])[0][0]
            if similarity > max_similarity:
                max_similarity = similarity
                best_keyword = row[col_keyword]
                best_label = row[output_column]

    return best_label, best_keyword


# Nhận diện nguyên liệu sử dụng từ khóa và phần trăm
def identify_materials(text, keywords_df):
    text = text.lower()
    materials_found = []
    for _, row in keywords_df.iterrows():
        keyword = row[col_keyword]
        if pd.isna(keyword):
            continue
        keyword = keyword.lower()
        material_name = row[col_material_name]
        matches = re.findall(r'(\d+%?)\s*' + re.escape(keyword), text)
        for match in matches:
            materials_found.append((match, material_name))
    return materials_found

def label_material(row, material_data, model):
    product_text = row[col_products].lower()  # Use the original 'Products' column
    keyword_positions = {k: product_text.find(k.lower()) for k in material_data[col_keyword].dropna().unique() if k.lower() in product_text}
    materials = identify_materials(product_text, material_data)
    material_labels = {}
    material_names_no_percentage = set()

    for match, material_name in materials:
        percentage_match = re.match(r'(\d+%?)', match)
        if percentage_match:
            percentage = percentage_match.group(1)
            material_labels[material_name] = percentage
            material_names_no_percentage.add(material_name)
        else:
            material_labels[material_name] = '100%'
            material_names_no_percentage.add(material_name)

    combined_labels = []
    for material, percentage in material_labels.items():
        combined_labels.append(f"{percentage} {material}")

    if combined_labels:
        return pd.Series([', '.join(combined_labels), ', '.join(material_names_no_percentage)], index=[col_material_name, col_material_name_no_percentage])
    else:
        material_label, material_keyword = label_using_word2vec(product_text, material_data, model, col_material_name)
        if material_label == "No label":
            material_names_no_percentage.add("No label")
            return pd.Series([material_label, material_label], index=[col_material_name, col_material_name_no_percentage])
        elif not re.search(r'\d+%?', material_label):
            material_label = f"100% {material_label}"
            material_names_no_percentage.add(material_label.split('100% ')[1])
        return pd.Series([material_label, ', '.join(material_names_no_percentage)], index=[col_material_name, col_material_name_no_percentage])

def determine_material_type(description):
    description_lower = description.lower()
    if any(keyword in description_lower for keyword in ["dệt thoi", "woven"]):
        return "Vải dệt thoi"
    elif any(keyword in description_lower for keyword in ["dệt kim", "knit"]):
        return "Vải dệt kim"
    elif any(keyword in description_lower for keyword in ["không dệt"]):
        return "Vải không dệt"
    else:
        return "Vải khác"

# Kiểm tra các chỉ số đặc biệt của sợi
def check_special_indicators(text):
    special_indicators = {
        'Denier': r'\b\d{1,3}\s*/\s*\d{1,3}[Dd]\b',  # Pattern for Denier, ví dụ: 20/22D
        'Dtex': r'\b\d{1,3}\s*[Dd][Tt][Ee][Xx]\b',   # Pattern for Dtex, ví dụ: 75Dtex
        'Tex': r'\b\d{1,3}\s*[Tt][Ee][Xx]\b',        # Pattern for Tex, ví dụ: 40Tex
        'Ne': r'\b[Nn][Ee]\s*\d{1,3}\b',             # Pattern for Ne, ví dụ: Ne 30
    }

    for label, indicator in special_indicators.items():
        match = re.search(indicator, text)
        if match:
            return match.group(0)  # Trả về chuỗi đã khớp, ví dụ: "20/22D"

    return "No Label"


# Extract physical properties using regex
def extract_physical_properties(row):
    properties = {
        col_Thickness: r'\b(dày|thickness)\s*\d+(\.\d+)?\s*(mm|mil)\b',
        col_Width: r'\b(rộng|width)\s*\d+(\.\d+)?(-\d+(\.\d+)?\s*)?(cm|m|inch|")\b',
        col_Weight_range: r'\b(trọng lượng\s*(không quá|≤|:|đ/l:)?\s*\d+(\.\d+)?(-\d+(\.\d+)?\s*)?(gsm|g/m2|gm/mtr|gms))\b',
        col_fabric_size: r'\b(khổ|khổ vải|khổ rộng)\s*[:;]?\s*\d+([\.,]\d+)?(\s*[-x/_]\s*\d+([\.,]\d+)?)*\s*(cm|m|inch|")?\b'
    }

    for col, pattern in properties.items():
        match = re.search(pattern, row[col_products].lower(), re.IGNORECASE)
        if match:
            row[col] = match.group(0)
        else:
            row[col] = "Not found"

    return row

def detect_new_percentage(text):
    pattern = r'\bmới\s*([\d]+%?)\b'
    match = re.search(pattern, text, re.IGNORECASE)
    if match:
        return f"Mới {match.group(1)}%"
    else:
        return "No Label"

def process_and_label_dataframe(df, model, keywords_dict):
    # Áp dụng chỉnh sửa chính tả cho cột Products
    df[col_corrected_products] = df[col_products].apply(correct_spelling_with_dict).astype(str)

    for col, (sheet_name, label_col) in keywords_dict.items():
        keywords = load_data(keywords_file_path, sheet_name=sheet_name)

        if col == col_material_name:
            # Áp dụng gán nhãn vật liệu
            df[[col_material_name, col_material_name_no_percentage]] = df.apply(
                lambda row: label_material(row, keywords, model), axis=1)
        else:
            # Áp dụng gán nhãn dựa trên Word2Vec, truyền vào supply_chain khi cần
            df[[col, 'Keyword_' + col]] = df.apply(
                lambda row: pd.Series(label_using_word2vec(
                    row[col_corrected_products], keywords, model, label_col,
                    row[col_supply_chain] if col == col_model_pattern else None)),
                axis=1)

    # Nhận diện sản phẩm mới và thêm vào cột "Sản phẩm mới"
    df['Sản phẩm mới'] = df[col_corrected_products].apply(detect_new_percentage)

    # Gán nhãn cho cột "Chi số sợi" bằng cách sử dụng hàm check_special_indicators
    df[col_yarn_count] = df[col_corrected_products].apply(check_special_indicators)

    # Cập nhật nhãn "Chuỗi cung ứng" nếu có "Chi số sợi" và ban đầu là "Nguyên liệu"
    def update_supply_chain(row):
        if row[col_yarn_count] != "No Label" and row[col_supply_chain].lower() == "nguyên liệu":
            return "Sợi"
        return row[col_supply_chain]

    df[col_supply_chain] = df.apply(update_supply_chain, axis=1)

    # Trích xuất các thuộc tính vật lý
    df = df.apply(extract_physical_properties, axis=1)

    return df

# Save labeled data back to Excel
def save_to_excel(file_path, df, keywords_dict):
    wb = load_workbook(file_path)
    sheet = wb.active

    # Lưu các cột đã có trong keywords_dict
    for idx, (col, _) in enumerate(keywords_dict.items()):
        col_index = sheet.max_column + 1
        sheet.cell(row=1, column=col_index, value=col).font = Font(bold=True)

        for row_idx, row in df.iterrows():
            sheet.cell(row=row_idx + 2, column=col_index, value=row[col])

    # Thêm cột Tên nguyên liệu và Phân bổ thành phần

    col_material_name_no_percentage_index = sheet.max_column + 1
    sheet.cell(row=1, column=col_material_name_no_percentage_index, value=col_material_name_no_percentage).font = Font(bold=True)

    for row_idx, row in df.iterrows():

        sheet.cell(row=row_idx + 2, column=col_material_name_no_percentage_index, value=row[col_material_name_no_percentage])

    # Lưu cột "Sản phẩm mới"
    col_san_pham_moi_index = sheet.max_column + 1
    sheet.cell(row=1, column=col_san_pham_moi_index, value='Sản phẩm mới').font = Font(bold=True)
    for row_idx, row in df.iterrows():
        sheet.cell(row=row_idx + 2, column=col_san_pham_moi_index, value=row['Sản phẩm mới'])

    # Lưu cột "Chi số sợi"
    col_yarn_count_index = sheet.max_column + 1
    sheet.cell(row=1, column=col_yarn_count_index, value=col_yarn_count).font = Font(bold=True)
    for row_idx, row in df.iterrows():
        sheet.cell(row=row_idx + 2, column=col_yarn_count_index, value=row[col_yarn_count])

    # Lưu lại các thuộc tính vật lý
    for idx, col in enumerate([col_Thickness, col_Width, col_Weight_range, col_fabric_size]):
        col_index = sheet.max_column + 1
        sheet.cell(row=1, column=col_index, value=col).font = Font(bold=True)

        for row_idx, row in df.iterrows():
            sheet.cell(row=row_idx + 2, column=col_index, value=row[col])

    output_file_path = file_path.replace('converted_date', 'Labeled_Word2Vec_Re')
    wb.save(output_file_path)
    wb.close()
    logging.info(f"Processed data saved to '{output_file_path}'")

# Main processing loop
def main():
    # Khởi tạo từ điển cho các cột và nhãn tương ứng
    keywords_dict = {
        col_supply_chain: (sheet_name_supply_chain, col_supply_chain),
        col_product_label: (sheet_name_product, col_product_label),
        col_material_name: (sheet_name_material, col_material_name),
        col_model_pattern: (sheet_name_model_pattern, col_model_pattern),
        col_function: (sheet_name_function, col_function),
        col_shape: (sheet_name_shape, col_shape),
        col_pattern: (sheet_name_pattern, col_pattern),
    }

    # Tạo danh sách các câu từ mô tả sản phẩm để huấn luyện Word2Vec
    sentences = []
    for file_path in import_file_paths:
        df = load_data(file_path, sheet_name='Sheet1')
        product_descriptions = df[col_products].apply(str).tolist()
        for desc in product_descriptions:
            sentences.append(desc.lower().split())

    # Huấn luyện mô hình Word2Vec trên mô tả sản phẩm
    model = train_word2vec(sentences)

    # Xử lý từng tệp và áp dụng gán nhãn
    for file_path in import_file_paths:
        df = load_data(file_path, sheet_name='Sheet1')
        df = process_and_label_dataframe(df, model, keywords_dict)
        save_to_excel(file_path, df, keywords_dict)

if __name__ == "__main__":
    main()